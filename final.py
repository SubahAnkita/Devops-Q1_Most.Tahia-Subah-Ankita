import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl

# Prompt user for the search query
search_query = input("Enter the search query: ")

# Set up the WebDriver (e.g., ChromeDriver)
driver = webdriver.Chrome()  # Make sure 'chromedriver' is in your PATH

# Open Google
driver.get('https://www.google.com')

# Find the search box and enter the user's query
search_box = driver.find_element(By.NAME, 'q')
search_box.send_keys(search_query)

# Give some time for auto-suggestions to appear
time.sleep(2)

# Find the parent <div> that contains the auto-suggestions
parent_div = driver.find_element(By.CLASS_NAME, 'erkvQe')

# Find all the <li> elements inside the parent <div>
suggestions = parent_div.find_elements(By.TAG_NAME, 'li')

# Extract and determine the longest and shortest strings
longestString = ""
shortestString = "ThisIsUsedToFindTheMinimumLengthString."
print("These are the suggested words: \n")
for suggestion in suggestions:
    text_element = suggestion.find_element(By.TAG_NAME, 'span')
    print(text_element.text)
    string = text_element.text
    if len(string) > len(longestString):
        longestString = string
    if len(string) < len(shortestString):
        shortestString = string  

print("-------------------------------------- \n")
print("shortest : ", shortestString)
print("longest : ", longestString)

# Close the WebDriver
driver.quit()

# Determine the current day of the week
current_day = datetime.datetime.today().strftime('%A')  # e.g., 'Friday'

# Load the workbook and select the correct sheet based on the day
workbook_path = r'Keywords.xlsx'
workbook = openpyxl.load_workbook(workbook_path)
sheet = workbook[current_day]  # This will dynamically select the correct sheet

# Find the next available row in each column
next_row_keyword = sheet.max_row + 1
next_row_longest = sheet.max_row + 1
next_row_shortest = sheet.max_row + 1

# Write the query, longest and shortest suggestions to the next available row in respective columns
sheet[f'A{next_row_keyword}'] = search_query
sheet[f'B{next_row_longest}'] = longestString
sheet[f'C{next_row_shortest}'] = shortestString

# Save the workbook
workbook.save(workbook_path)
